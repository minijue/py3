import openpyxl as xl

if __name__ == '__main__':
    print("Please use me as a module.")


class OneClass:
    def __init__(self, name):
        self.name = name
        self.__students = {}
        self.norm = None
        self.exam = None
        self.scale = '0%'
        self.scoreType = 'score'

    def AddStudentScore(self, name, nscore, escore):
        self.__students[name] = (nscore, escore)

    def GetStudents(self):
        return self.__students


class ExcelFile():
    def __init__(self, path):
        self.wb = xl.load_workbook(path, data_only=True)
        self.sheetnames = self.wb.get_sheet_names()

    def getClassNames(self):
        return list(self.sheetnames)

    def getClasses(self):
        for name in self.sheetnames:
            cs = OneClass(name)  # 获取班级名
            ws = self.wb[name]
            # 获取平时成绩比例
            d1 = self.wb[name]['D1']
            cs.scale = d1.value[-3:]

            nColumn, eColumn = 0, 0

            # 找到平时成绩和考核成绩所在列
            for i in range(5, 50):
                c = ws.cell(row=1, column=i)
                if c.value == u'平时成绩总成绩':
                    nColumn = i

                if c.value == u'期末考核成绩':
                    eColumn = i

                if nColumn > 0 and eColumn > 0:
                    break

            # 获取平时成绩和考核成绩
            for j in range(3, 60):
                na = ws.cell(row=j, column=2)
                cn = ws.cell(row=j, column=nColumn)
                ce = ws.cell(row=j, column=eColumn)

                if na.value is None:
                    break

                # 根据单元格值的类型判断成绩类型
                if ce.value in (u'优秀', u'良好', u'中等', u'及格', u'不及格'):
                    cs.scoreType = u'fiveTypescore'
                elif ce.value in (u'合格', u'不合格'):
                    cs.scoreType = u'twoTypescore'

                cs.AddStudentScore(na.value, cn.value, ce.value)

            yield cs
